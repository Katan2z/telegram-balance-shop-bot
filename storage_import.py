import argparse
import json
import re
from pathlib import Path

import openpyxl


DURATION_RE = re.compile(
    r"(?P<value>\d+(?:[.,]\d+)?)\s*"
    r"(?P<unit>д(?:ень|ня|ней)|сут(?:ки|ок)?|час(?:а|ов)?|ч\.?|мин(?:ута|уты|ут)?)",
    re.IGNORECASE,
)


def clean_cell(value) -> str:
    return "" if value is None else str(value).strip()


def parse_duration_hours(value) -> float | None:
    """Read the first compound duration, such as '1 час 30 минут' or '5 дней'."""
    text = clean_cell(value).lower()
    if not text or text in {"-", "—"}:
        return None

    matches = list(DURATION_RE.finditer(text))
    if not matches:
        return None

    first = matches[0]
    number = float(first.group("value").replace(",", "."))
    unit = first.group("unit")
    if unit.startswith("д") or unit.startswith("сут"):
        total = number * 24
    elif unit.startswith("мин"):
        total = number / 60
    else:
        total = number

    if len(matches) > 1 and not unit.startswith(("д", "сут", "мин")):
        second = matches[1]
        gap = text[first.end() : second.start()]
        if second.group("unit").startswith("мин") and not re.search(r"[;()/\n]", gap):
            total += float(second.group("value").replace(",", ".")) / 60

    return total if total > 0 else None


def is_section_row(values: list[str]) -> bool:
    return values[7].lower() == "маркировка" and values[8].lower() == "место маркировки"


def item_from_row(row_number: int, values: list[str]) -> dict:
    prep_raw, defrost_raw, storage_raw = values[1], values[3], values[4]
    production_label = values[7]
    return {
        "name": values[0],
        "prepHours": parse_duration_hours(prep_raw),
        "defrostType": values[2],
        "defrostHours": parse_duration_hours(defrost_raw),
        "storageHours": parse_duration_hours(storage_raw),
        "prepRaw": prep_raw,
        "defrostRaw": defrost_raw,
        "storageRaw": storage_raw,
        "prepLabel": values[5],
        "prepPlace": values[6],
        "productionHours": parse_duration_hours(production_label),
        "productionLabel": production_label,
        "productionPlace": values[8],
        "sourceRow": row_number,
    }


def read_storage_items(path: str | Path) -> list[dict]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    items = []
    for row_number in range(6, sheet.max_row + 1):
        values = [clean_cell(sheet.cell(row_number, column).value) for column in range(1, 10)]
        if values[0].startswith("П = время начала"):
            break
        if not values[0] or is_section_row(values):
            continue
        items.append(item_from_row(row_number, values))
    workbook.close()
    return items


def write_storage_js(items: list[dict], output_path: str | Path) -> None:
    payload = json.dumps(items, ensure_ascii=False, indent=2)
    Path(output_path).write_text(f"window.STORAGE_ITEMS = {payload};\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Import BK8 storage times from Excel")
    parser.add_argument("xlsx", help="Source .xlsx file")
    parser.add_argument("--output", default="docs/storage-times.js")
    args = parser.parse_args()
    items = read_storage_items(args.xlsx)
    if not items:
        raise RuntimeError("No storage items found")
    write_storage_js(items, args.output)
    print(f"Imported {len(items)} storage items into {args.output}")


if __name__ == "__main__":
    main()
