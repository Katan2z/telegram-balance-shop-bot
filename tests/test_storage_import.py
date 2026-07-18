import tempfile
import unittest
from pathlib import Path

import openpyxl

from storage_import import parse_duration_hours, read_storage_items


class StorageImportTests(unittest.TestCase):
    def test_parses_compound_duration(self):
        self.assertEqual(parse_duration_hours("1 час 30 минут"), 1.5)
        self.assertEqual(parse_duration_hours("ТД = 5 дней"), 120)
        self.assertEqual(parse_duration_hours("70 часов 30 минут"), 70.5)

    def test_reads_products_and_skips_section_rows(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.cell(6, 1, "Продукт")
        sheet.cell(6, 2, "54 часа")
        sheet.cell(6, 4, "6 часов")
        sheet.cell(6, 5, "48 часов")
        sheet.cell(6, 8, "ТД = 4 часа")
        sheet.cell(7, 1, "Раздел")
        sheet.cell(7, 8, "маркировка")
        sheet.cell(7, 9, "место маркировки")
        sheet.cell(8, 1, "П = время начала разморозки")

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "storage.xlsx"
            workbook.save(path)
            items = read_storage_items(path)

        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["name"], "Продукт")
        self.assertEqual(items[0]["prepHours"], 54)
        self.assertEqual(items[0]["defrostHours"], 6)
        self.assertEqual(items[0]["storageHours"], 48)


if __name__ == "__main__":
    unittest.main()
